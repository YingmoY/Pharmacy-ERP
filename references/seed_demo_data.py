import os
import subprocess
from datetime import date, datetime, timedelta

import openpyxl


ROOT = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(ROOT, "references", "药品信息.xlsx")
BASE_DATE = date(2026, 6, 14)
DATE_STR = BASE_DATE.strftime("%Y%m%d")
REQUEST_ID = "codex-demo-seed"


def q(value):
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return str(value)
    s = str(value).replace("\x00", "").strip()
    return "'" + s.replace("'", "''") + "'"


def pick(row, headers, *names):
    for name in names:
        idx = headers.get(name)
        if idx is not None and row[idx] not in (None, ""):
            return str(row[idx]).strip()
    return ""


def load_drugs(limit=20):
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    headers = {name: i for i, name in enumerate(header)}
    drugs = []
    seen = set()
    seen_barcodes = set()
    for row in rows:
        drug_code = pick(row, headers, "med_list_codg", "drugstdcode", "id")[:50]
        name = pick(row, headers, "drug_name", "reg_name", "drug_genname")
        spec = pick(row, headers, "drug_size", "drug_spec", "reg_spec")
        manufacturer = pick(row, headers, "prodentp_name", "co_name", "subpck_fcty")
        if not drug_code or not name or not spec or not manufacturer or drug_code in seen:
            continue
        seen.add(drug_code)
        i = len(drugs) + 1
        trade = pick(row, headers, "drug_prodname")
        if trade == "无":
            trade = ""
        barcode = pick(row, headers, "drugstdcode")[:100] or None
        if barcode in seen_barcodes:
            barcode = None
        if barcode:
            seen_barcodes.add(barcode)
        drugs.append(
            {
                "code": drug_code,
                "common_name": name[:100],
                "trade_name": trade[:100],
                "specification": spec[:50],
                "dosage_form": pick(row, headers, "drug_dosform", "reg_dosform")[:50],
                "manufacturer": manufacturer[:100],
                "approval_number": pick(row, headers, "aprvno", "nat_drug_no")[:50],
                "barcode": barcode,
                "unit": pick(row, headers, "min_salunt", "min_unt", "min_pacunt")[:20] or "盒",
                "is_prescription": i % 4 == 0,
                "is_medicare": True,
                "purchase_price": round(8 + i * 1.7, 2),
                "retail_price": round(13.8 + i * 2.35, 2),
                "storage_condition": "常温、避光、干燥处保存",
                "remark": "示例数据，药品基础信息来自 references/药品信息.xlsx",
            }
        )
        if len(drugs) >= limit:
            break
    return drugs


def num(ts):
    return f"860000{DATE_STR}{ts:06d}"


def ts(days_offset, hour, minute=0):
    d = datetime.combine(BASE_DATE + timedelta(days=days_offset), datetime.min.time())
    return d.replace(hour=hour, minute=minute)


def build_sql():
    drugs = load_drugs()
    sql = [
        "BEGIN;",
        "SET search_path = public;",
        "",
        "-- Clean only this deterministic demo seed.",
        "DELETE FROM report_export_task WHERE task_id LIKE 'RPT-DEMO-20260614-%';",
        "DELETE FROM notification WHERE business_id LIKE 'DEMO-20260614-%';",
        "DELETE FROM audit_event WHERE related_id LIKE 'DEMO-20260614-%';",
        "DELETE FROM security_event WHERE description LIKE '%Codex demo seed%';",
        "DELETE FROM inventory_adjustment WHERE adjust_no LIKE 'ADJ-20260614-9%';",
        "DELETE FROM inventory_task_detail WHERE task_id IN (SELECT id FROM inventory_task WHERE task_no LIKE 'INV-20260614-9%');",
        "DELETE FROM inventory_task WHERE task_no LIKE 'INV-20260614-9%';",
        "DELETE FROM scan_task_detail WHERE task_id IN (SELECT id FROM scan_task WHERE task_no LIKE 'SCAN-20260614-9%');",
        "DELETE FROM scan_task WHERE task_no LIKE 'SCAN-20260614-9%';",
        "DELETE FROM trace_reservation WHERE reservation_no LIKE 'RSV-20260614-9%';",
        f"DELETE FROM drug_trace_log WHERE request_id = {q(REQUEST_ID)} OR related_no LIKE '%20260614-9%';",
        "DELETE FROM audit_review WHERE review_no LIKE 'REV-20260614-9%';",
        "DELETE FROM sales_order_item WHERE order_id IN (SELECT id FROM sales_order WHERE order_no LIKE 'SO-20260614-9%');",
        "DELETE FROM sales_order WHERE order_no LIKE 'SO-20260614-9%';",
        "DELETE FROM drug_trace_inventory WHERE trace_code LIKE '86000020260614%';",
        "DELETE FROM inbound_order_detail WHERE order_id IN (SELECT id FROM inbound_order WHERE order_no LIKE 'IN-20260614-9%');",
        "DELETE FROM inbound_order WHERE order_no LIKE 'IN-20260614-9%';",
        "",
        "-- Stable demo users and role bindings.",
    ]

    users = [
        ("demo_manager", "示例店长", "STORE_MANAGER"),
        ("demo_pharmacist", "示例药师", "PHARMACIST"),
        ("demo_cashier", "示例收银员", "CASHIER"),
        ("demo_warehouse", "示例仓库员", "WAREHOUSE"),
    ]
    for username, real_name, _ in users:
        sql.append(
            "INSERT INTO sys_user (username,password_hash,real_name,status,phone,email,remark) VALUES "
            f"({q(username)}, {q('$2a$10$demoSeedHashOnlyForLocalData')}, {q(real_name)}, 1, "
            f"{q('1380000' + str(9000 + len(sql))[-4:])}, {q(username + '@demo.local')}, {q('Codex demo seed user')}) "
            "ON CONFLICT (username) DO UPDATE SET real_name=EXCLUDED.real_name,status=1,updated_at=CURRENT_TIMESTAMP;"
        )
    for username, _, role_code in users:
        sql.append(
            "INSERT INTO sys_user_role (user_id, role_id) "
            f"SELECT u.id, r.id FROM sys_user u, sys_role r WHERE u.username={q(username)} AND r.code={q(role_code)} "
            "ON CONFLICT (user_id, role_id) DO NOTHING;"
        )

    sql.append("")
    sql.append("-- Suppliers and locations.")
    suppliers = [
        ("SUP-9001", "华东医药商业配送有限公司", "陈洁", "021-55559001"),
        ("SUP-9002", "国药控股示例医药有限公司", "李明", "021-55559002"),
        ("SUP-9003", "九州通示例供应链有限公司", "王珊", "021-55559003"),
    ]
    for code, name, contact, phone in suppliers:
        sql.append(
            "INSERT INTO supplier (supplier_code,name,contact_name,contact_phone,license_no,address,status,remark) VALUES "
            f"({q(code)},{q(name)},{q(contact)},{q(phone)},{q('LIC-' + code)},{q('上海市浦东新区示例路' + code[-1] + '号')},1,{q('Codex demo seed supplier')}) "
            "ON CONFLICT (supplier_code) DO UPDATE SET name=EXCLUDED.name,status=1,updated_at=CURRENT_TIMESTAMP;"
        )
    for area, shelf_count in [("OTC", 4), ("RX", 4), ("COLD", 2)]:
        for n in range(1, shelf_count + 1):
            code = f"{area}-A{n:02d}-L1-P01"
            sql.append(
                "INSERT INTO location_info (location_code,location_name,area,shelf,layer,position,capacity,status,remark) VALUES "
                f"({q(code)},{q(area + ' 示例货位 ' + str(n))},{q(area)},{q('A' + str(n).zfill(2))},{q('L1')},{q('P01')},120,1,{q('Codex demo seed location')}) "
                "ON CONFLICT (location_code) DO UPDATE SET location_name=EXCLUDED.location_name,status=1,updated_at=CURRENT_TIMESTAMP;"
            )

    sql.append("")
    sql.append("-- Real drug master data from references/药品信息.xlsx.")
    for d in drugs:
        sql.append(
            "INSERT INTO drug_info (drug_code,common_name,trade_name,specification,dosage_form,manufacturer,approval_number,"
            "is_prescription,is_medicare,status,barcode,unit,retail_price,purchase_price,storage_condition,remark) VALUES "
            f"({q(d['code'])},{q(d['common_name'])},{q(d['trade_name'])},{q(d['specification'])},{q(d['dosage_form'])},"
            f"{q(d['manufacturer'])},{q(d['approval_number'])},{q(d['is_prescription'])},{q(d['is_medicare'])},1,"
            f"{q(d['barcode'])},{q(d['unit'])},{d['retail_price']},{d['purchase_price']},{q(d['storage_condition'])},{q(d['remark'])}) "
            "ON CONFLICT (drug_code) DO UPDATE SET common_name=EXCLUDED.common_name,trade_name=EXCLUDED.trade_name,"
            "specification=EXCLUDED.specification,dosage_form=EXCLUDED.dosage_form,manufacturer=EXCLUDED.manufacturer,"
            "approval_number=EXCLUDED.approval_number,is_prescription=EXCLUDED.is_prescription,is_medicare=EXCLUDED.is_medicare,"
            "barcode=EXCLUDED.barcode,unit=EXCLUDED.unit,retail_price=EXCLUDED.retail_price,purchase_price=EXCLUDED.purchase_price,"
            "storage_condition=EXCLUDED.storage_condition,remark=EXCLUDED.remark,status=1,updated_at=CURRENT_TIMESTAMP;"
        )

    inbound_orders = []
    trace_rows = []
    trace_seq = 1
    for o in range(4):
        order_no = f"IN-{DATE_STR}-{9001 + o:04d}"
        supplier = suppliers[o % len(suppliers)][0]
        created = ts(-20 + o * 3, 9, 10)
        completed = created + timedelta(hours=2)
        inbound_orders.append(order_no)
        detail_drugs = drugs[o * 5 : o * 5 + 5]
        total = sum(d["purchase_price"] * 4 for d in detail_drugs)
        sql.append(
            "INSERT INTO inbound_order (order_no,invoice_no,operator_id,creator_id,status,remark,supplier_id,total_amount,submitted_at,completed_at,created_at,updated_at) VALUES "
            f"({q(order_no)},{q('FP' + DATE_STR + str(9001 + o))},"
            "(SELECT id FROM sys_user WHERE username='demo_warehouse'),(SELECT id FROM sys_user WHERE username='demo_manager'),"
            f"'COMPLETED',{q('Codex demo seed completed inbound order')},(SELECT id FROM supplier WHERE supplier_code={q(supplier)}),"
            f"{total:.2f},{q(created + timedelta(minutes=20))},{q(completed)},{q(created)},{q(completed)});"
        )
        for j, d in enumerate(detail_drugs):
            batch = f"B{DATE_STR[-4:]}{o + 1}{j + 1:02d}"
            expire = BASE_DATE + timedelta(days=420 + j * 30 + o * 15)
            amount = d["purchase_price"] * 4
            sql.append(
                "INSERT INTO inbound_order_detail (order_id,drug_id,batch_number,expire_date,planned_qty,confirmed_qty,unit_price,amount,remark,created_at,updated_at) VALUES "
                f"((SELECT id FROM inbound_order WHERE order_no={q(order_no)}),(SELECT id FROM drug_info WHERE drug_code={q(d['code'])}),"
                f"{q(batch)},{q(expire)},4,4,{d['purchase_price']},{amount:.2f},{q('Codex demo seed inbound detail')},{q(created)},{q(completed)});"
            )
            for k in range(4):
                trace_code = num(trace_seq)
                trace_seq += 1
                loc_area = "RX" if d["is_prescription"] else "OTC"
                loc = f"{loc_area}-A{((j + k) % 4) + 1:02d}-L1-P01"
                trace_rows.append((trace_code, d, order_no, batch, expire, loc))

    sold_traces = trace_rows[:12]
    refunded_traces = trace_rows[12:14]
    lost_trace = trace_rows[14]
    relocation_trace = trace_rows[15]

    for idx, (trace_code, d, order_no, batch, expire, loc) in enumerate(trace_rows):
        status = "IN_STOCK"
        sold_at = "NULL"
        last_action = "SHELVED"
        if (trace_code, d, order_no, batch, expire, loc) in sold_traces:
            status = "SOLD"
            sold_at = q(ts(-4 + idx % 3, 15, 20))
            last_action = "SOLD"
        if trace_code == lost_trace[0]:
            status = "LOST"
            last_action = "LOSS_CONFIRMED"
        if trace_code == relocation_trace[0]:
            loc = "OTC-A04-L1-P01"
            last_action = "RELOCATED"
        sql.append(
            "INSERT INTO drug_trace_inventory (trace_code,drug_id,batch_number,expire_date,location_id,status,inbound_order_id,inbound_detail_id,sold_at,last_action,created_at,updated_at) VALUES "
            f"({q(trace_code)},(SELECT id FROM drug_info WHERE drug_code={q(d['code'])}),{q(batch)},{q(expire)},"
            f"(SELECT id FROM location_info WHERE location_code={q(loc)}),{q(status)},"
            f"(SELECT id FROM inbound_order WHERE order_no={q(order_no)}),"
            f"(SELECT id FROM inbound_order_detail WHERE order_id=(SELECT id FROM inbound_order WHERE order_no={q(order_no)}) AND drug_id=(SELECT id FROM drug_info WHERE drug_code={q(d['code'])}) AND batch_number={q(batch)}),"
            f"{sold_at},{q(last_action)},{q(ts(-18, 11))},{q(ts(-3, 16))});"
        )
        for action, from_status, to_status in [("INBOUND_CONFIRM", "", "PENDING"), ("SHELVING", "PENDING", "IN_STOCK")]:
            sql.append(
                "INSERT INTO drug_trace_log (trace_code,action_type,from_status,to_status,operator_id,related_no,remark,drug_id,request_id,to_location_id,created_at,updated_at) VALUES "
                f"({q(trace_code)},{q(action)},{q(from_status)},{q(to_status)},(SELECT id FROM sys_user WHERE username='demo_warehouse'),"
                f"{q(order_no)},{q('Codex demo seed trace log')},(SELECT id FROM drug_info WHERE drug_code={q(d['code'])}),{q(REQUEST_ID)},"
                f"(SELECT location_id FROM drug_trace_inventory WHERE trace_code={q(trace_code)}),{q(ts(-18, 11))},{q(ts(-18, 11))});"
            )

    sale_groups = [sold_traces[i : i + 2] for i in range(0, 12, 2)]
    for i, group in enumerate(sale_groups):
        order_no = f"SO-{DATE_STR}-{9001 + i:04d}"
        created = ts(-6 + i, 14, 5)
        total = sum(g[1]["retail_price"] for g in group)
        is_rx = any(g[1]["is_prescription"] for g in group)
        sql.append(
            "INSERT INTO sales_order (order_no,cashier_id,total_amount,medicare_amount,personal_amount,need_audit,need_medicare,status,"
            "customer_name,is_prescription,discount_amount,actual_amount,payment_method,paid_at,refund_amount,remark,created_at,updated_at) VALUES "
            f"({q(order_no)},(SELECT id FROM sys_user WHERE username='demo_cashier'),{total:.2f},0,{total:.2f},{q(is_rx)},FALSE,'COMPLETED',"
            f"{q('示例顾客' + str(i + 1))},{q(is_rx)},0,{total:.2f},{q('CASH')},{q(created + timedelta(minutes=8))},0,"
            f"{q('Codex demo seed completed sale')},{q(created)},{q(created + timedelta(minutes=8))});"
        )
        for j, (trace_code, d, *_rest) in enumerate(group):
            sql.append(
                "INSERT INTO sales_order_item (order_id,drug_id,trace_code,price,quantity,subtotal_amount,remark,refund_status,refund_amount,created_at,updated_at) VALUES "
                f"((SELECT id FROM sales_order WHERE order_no={q(order_no)}),(SELECT id FROM drug_info WHERE drug_code={q(d['code'])}),"
                f"{q(trace_code)},{d['retail_price']},1,{d['retail_price']},{q('Codex demo seed sale item')},'NONE',0,{q(created)},{q(created)});"
            )
            rsv_no = f"RSV-{DATE_STR}-{9001 + i * 2 + j:04d}"
            sql.append(
                "INSERT INTO trace_reservation (reservation_no,sales_order_id,sales_order_item_id,trace_code,drug_id,reserved_by,status,reserved_at,confirmed_at,expire_at,remark,created_at,updated_at) VALUES "
                f"({q(rsv_no)},(SELECT id FROM sales_order WHERE order_no={q(order_no)}),"
                f"(SELECT id FROM sales_order_item WHERE order_id=(SELECT id FROM sales_order WHERE order_no={q(order_no)}) AND trace_code={q(trace_code)}),"
                f"{q(trace_code)},(SELECT id FROM drug_info WHERE drug_code={q(d['code'])}),(SELECT id FROM sys_user WHERE username='demo_cashier'),"
                f"'CONSUMED',{q(created)},{q(created + timedelta(minutes=8))},{q(created + timedelta(hours=1))},{q('Codex demo seed consumed reservation')},{q(created)},{q(created)});"
            )
            sql.append(
                "INSERT INTO drug_trace_log (trace_code,action_type,from_status,to_status,operator_id,related_no,remark,drug_id,order_id,order_item_id,request_id,created_at,updated_at) VALUES "
                f"({q(trace_code)},'SALE','IN_STOCK','SOLD',(SELECT id FROM sys_user WHERE username='demo_cashier'),{q(order_no)},"
                f"{q('Codex demo seed sold')},(SELECT id FROM drug_info WHERE drug_code={q(d['code'])}),(SELECT id FROM sales_order WHERE order_no={q(order_no)}),"
                f"(SELECT id FROM sales_order_item WHERE order_id=(SELECT id FROM sales_order WHERE order_no={q(order_no)}) AND trace_code={q(trace_code)}),"
                f"{q(REQUEST_ID)},{q(created + timedelta(minutes=8))},{q(created + timedelta(minutes=8))});"
            )
        if is_rx:
            rev_no = f"REV-{DATE_STR}-{9001 + i:04d}"
            sql.append(
                "INSERT INTO audit_review (order_id,review_no,pharmacist_id,status,comment,reviewed_at,submitter_id,submitted_at,review_opinion,created_at,updated_at) VALUES "
                f"((SELECT id FROM sales_order WHERE order_no={q(order_no)}),{q(rev_no)},(SELECT id FROM sys_user WHERE username='demo_pharmacist'),"
                f"'APPROVED',{q('处方信息与用药风险核对通过')},{q(created + timedelta(minutes=5))},(SELECT id FROM sys_user WHERE username='demo_cashier'),"
                f"{q(created + timedelta(minutes=1))},{q('Codex demo seed approved review')},{q(created)},{q(created + timedelta(minutes=5))});"
            )

    refund_order = f"SO-{DATE_STR}-9007"
    created = ts(-2, 16, 10)
    total = sum(g[1]["retail_price"] for g in refunded_traces)
    sql.append(
        "INSERT INTO sales_order (order_no,cashier_id,total_amount,medicare_amount,personal_amount,need_audit,need_medicare,status,"
        "customer_name,is_prescription,discount_amount,actual_amount,payment_method,paid_at,refunded_at,refund_amount,refund_reason,remark,created_at,updated_at) VALUES "
        f"({q(refund_order)},(SELECT id FROM sys_user WHERE username='demo_cashier'),{total:.2f},0,{total:.2f},FALSE,FALSE,'REFUNDED',"
        f"{q('示例退货顾客')},FALSE,0,{total:.2f},{q('WECHAT')},{q(created + timedelta(minutes=5))},{q(created + timedelta(hours=2))},"
        f"{total:.2f},{q('顾客全单退货')},{q('Codex demo seed refunded sale')},{q(created)},{q(created + timedelta(hours=2))});"
    )
    for j, (trace_code, d, *_rest) in enumerate(refunded_traces):
        sql.append(
            "INSERT INTO sales_order_item (order_id,drug_id,trace_code,price,quantity,subtotal_amount,remark,refund_status,refund_amount,refunded_at,refund_reason,refund_operator_id,created_at,updated_at) VALUES "
            f"((SELECT id FROM sales_order WHERE order_no={q(refund_order)}),(SELECT id FROM drug_info WHERE drug_code={q(d['code'])}),"
            f"{q(trace_code)},{d['retail_price']},1,{d['retail_price']},{q('Codex demo seed refunded sale item')},'REFUNDED',{d['retail_price']},"
            f"{q(created + timedelta(hours=2))},{q('顾客全单退货')},(SELECT id FROM sys_user WHERE username='demo_cashier'),{q(created)},{q(created + timedelta(hours=2))});"
        )
        rsv_no = f"RSV-{DATE_STR}-{9013 + j:04d}"
        sql.append(
            "INSERT INTO trace_reservation (reservation_no,sales_order_id,sales_order_item_id,trace_code,drug_id,reserved_by,status,reserved_at,confirmed_at,expire_at,remark,created_at,updated_at) VALUES "
            f"({q(rsv_no)},(SELECT id FROM sales_order WHERE order_no={q(refund_order)}),"
            f"(SELECT id FROM sales_order_item WHERE order_id=(SELECT id FROM sales_order WHERE order_no={q(refund_order)}) AND trace_code={q(trace_code)}),"
            f"{q(trace_code)},(SELECT id FROM drug_info WHERE drug_code={q(d['code'])}),(SELECT id FROM sys_user WHERE username='demo_cashier'),"
            f"'CONSUMED',{q(created)},{q(created + timedelta(minutes=5))},{q(created + timedelta(hours=1))},{q('Codex demo seed consumed then refunded')},{q(created)},{q(created)});"
        )
        sql.append(f"UPDATE drug_trace_inventory SET status='IN_STOCK', sold_at=NULL, last_action='REFUNDED', updated_at={q(created + timedelta(hours=2))} WHERE trace_code={q(trace_code)};")
        sql.append(
            "INSERT INTO drug_trace_log (trace_code,action_type,from_status,to_status,operator_id,related_no,remark,drug_id,order_id,order_item_id,request_id,created_at,updated_at) VALUES "
            f"({q(trace_code)},'REFUND','SOLD','IN_STOCK',(SELECT id FROM sys_user WHERE username='demo_cashier'),{q(refund_order)},"
            f"{q('Codex demo seed refund')},(SELECT id FROM drug_info WHERE drug_code={q(d['code'])}),(SELECT id FROM sales_order WHERE order_no={q(refund_order)}),"
            f"(SELECT id FROM sales_order_item WHERE order_id=(SELECT id FROM sales_order WHERE order_no={q(refund_order)}) AND trace_code={q(trace_code)}),"
            f"{q(REQUEST_ID)},{q(created + timedelta(hours=2))},{q(created + timedelta(hours=2))});"
        )

    rx_trace_code, rx_drug, *_ = relocation_trace
    rx_order = f"SO-{DATE_STR}-9008"
    created = ts(0, 10, 20)
    sql.append(
        "INSERT INTO sales_order (order_no,cashier_id,total_amount,medicare_amount,personal_amount,need_audit,need_medicare,status,"
        "customer_name,is_prescription,discount_amount,actual_amount,payment_method,paid_at,refund_amount,remark,created_at,updated_at) VALUES "
        f"({q(rx_order)},(SELECT id FROM sys_user WHERE username='demo_cashier'),{rx_drug['retail_price']:.2f},0,{rx_drug['retail_price']:.2f},TRUE,FALSE,'COMPLETED',"
        f"{q('示例处方顾客')},TRUE,0,{rx_drug['retail_price']:.2f},{q('ALIPAY')},{q(created + timedelta(minutes=12))},0,"
        f"{q('Codex demo seed approved prescription sale')},{q(created)},{q(created + timedelta(minutes=12))});"
    )
    sql.append(
        "INSERT INTO sales_order_item (order_id,drug_id,trace_code,price,quantity,subtotal_amount,remark,refund_status,refund_amount,created_at,updated_at) VALUES "
        f"((SELECT id FROM sales_order WHERE order_no={q(rx_order)}),(SELECT id FROM drug_info WHERE drug_code={q(rx_drug['code'])}),"
        f"{q(rx_trace_code)},{rx_drug['retail_price']},1,{rx_drug['retail_price']},{q('Codex demo seed prescription sale item')},'NONE',0,{q(created)},{q(created)});"
    )
    sql.append(f"UPDATE drug_trace_inventory SET status='SOLD', sold_at={q(created + timedelta(minutes=12))}, last_action='SOLD', updated_at={q(created + timedelta(minutes=12))} WHERE trace_code={q(rx_trace_code)};")
    sql.append(
        "INSERT INTO trace_reservation (reservation_no,sales_order_id,sales_order_item_id,trace_code,drug_id,reserved_by,status,reserved_at,confirmed_at,expire_at,remark,created_at,updated_at) VALUES "
        f"({q('RSV-' + DATE_STR + '-9015')},(SELECT id FROM sales_order WHERE order_no={q(rx_order)}),"
        f"(SELECT id FROM sales_order_item WHERE order_id=(SELECT id FROM sales_order WHERE order_no={q(rx_order)}) AND trace_code={q(rx_trace_code)}),"
        f"{q(rx_trace_code)},(SELECT id FROM drug_info WHERE drug_code={q(rx_drug['code'])}),(SELECT id FROM sys_user WHERE username='demo_cashier'),"
        f"'CONSUMED',{q(created)},{q(created + timedelta(minutes=12))},{q(created + timedelta(hours=1))},{q('Codex demo seed consumed prescription reservation')},{q(created)},{q(created)});"
    )
    sql.append(
        "INSERT INTO audit_review (order_id,review_no,pharmacist_id,status,comment,reviewed_at,submitter_id,submitted_at,review_opinion,created_at,updated_at) VALUES "
        f"((SELECT id FROM sales_order WHERE order_no={q(rx_order)}),{q('REV-' + DATE_STR + '-9001')},(SELECT id FROM sys_user WHERE username='demo_pharmacist'),"
        f"'APPROVED',{q('处方药销售审核通过')},{q(created + timedelta(minutes=8))},(SELECT id FROM sys_user WHERE username='demo_cashier'),"
        f"{q(created + timedelta(minutes=2))},{q('Codex demo seed approved prescription review')},{q(created)},{q(created + timedelta(minutes=8))});"
    )
    sql.append(
        "INSERT INTO drug_trace_log (trace_code,action_type,from_status,to_status,operator_id,related_no,remark,drug_id,order_id,order_item_id,request_id,created_at,updated_at) VALUES "
        f"({q(rx_trace_code)},'SALE','IN_STOCK','SOLD',(SELECT id FROM sys_user WHERE username='demo_cashier'),{q(rx_order)},"
        f"{q('Codex demo seed prescription sold')},(SELECT id FROM drug_info WHERE drug_code={q(rx_drug['code'])}),(SELECT id FROM sales_order WHERE order_no={q(rx_order)}),"
        f"(SELECT id FROM sales_order_item WHERE order_id=(SELECT id FROM sales_order WHERE order_no={q(rx_order)}) AND trace_code={q(rx_trace_code)}),"
        f"{q(REQUEST_ID)},{q(created + timedelta(minutes=12))},{q(created + timedelta(minutes=12))});"
    )

    sql.append("")
    sql.append("-- Completed scan tasks.")
    for i, order_no in enumerate(inbound_orders):
        task_no = f"SCAN-{DATE_STR}-{9001 + i:04d}"
        started = ts(-19 + i * 3, 10, 0)
        sql.append(
            "INSERT INTO scan_task (task_no,task_type,related_id,operator_id,status,start_time,end_time,remark,created_at,updated_at) VALUES "
            f"({q(task_no)},'INBOUND',(SELECT id FROM inbound_order WHERE order_no={q(order_no)}),(SELECT id FROM sys_user WHERE username='demo_warehouse'),"
            f"'COMPLETED',{q(started)},{q(started + timedelta(hours=1))},{q('Codex demo seed inbound scan task')},{q(started)},{q(started + timedelta(hours=1))});"
        )
        for trace_code, _d, t_order_no, *_rest in [r for r in trace_rows if r[2] == order_no]:
            sql.append(
                "INSERT INTO scan_task_detail (task_id,trace_code,location_code,scan_result,scan_time,created_at,updated_at) VALUES "
                f"((SELECT id FROM scan_task WHERE task_no={q(task_no)}),{q(trace_code)},NULL,'SUCCESS',{q(started + timedelta(minutes=15))},{q(started)},{q(started)});"
            )

    inv_task_no = f"INV-{DATE_STR}-9001"
    start = ts(-1, 9, 0)
    sql.append(
        "INSERT INTO inventory_task (task_no,scope_type,scope_value,creator_id,assignee_id,status,start_time,end_time,remark,created_at,updated_at) VALUES "
        f"({q(inv_task_no)},'AREA','OTC',(SELECT id FROM sys_user WHERE username='demo_manager'),(SELECT id FROM sys_user WHERE username='demo_warehouse'),"
        f"'COMPLETED',{q(start)},{q(start + timedelta(hours=2))},{q('Codex demo seed completed inventory task')},{q(start)},{q(start + timedelta(hours=2))});"
    )
    for trace_code, _d, *_rest in trace_rows[20:28]:
        sql.append(
            "INSERT INTO inventory_task_detail (task_id,location_id,trace_code,discrepancy_type,scanned_location_id,system_location_id,operator_id,scanned_at,created_at,updated_at) VALUES "
            f"((SELECT id FROM inventory_task WHERE task_no={q(inv_task_no)}),(SELECT location_id FROM drug_trace_inventory WHERE trace_code={q(trace_code)}),"
            f"{q(trace_code)},'NORMAL',(SELECT location_id FROM drug_trace_inventory WHERE trace_code={q(trace_code)}),"
            f"(SELECT location_id FROM drug_trace_inventory WHERE trace_code={q(trace_code)}),(SELECT id FROM sys_user WHERE username='demo_warehouse'),"
            f"{q(start + timedelta(minutes=30))},{q(start)},{q(start)});"
        )

    adj_no = f"ADJ-{DATE_STR}-9001"
    sql.append(
        "INSERT INTO inventory_adjustment (adjust_no,adjust_type,trace_code,drug_id,from_location_id,to_location_id,before_status,after_status,reason,operator_id,related_task_id,status,created_at,updated_at) VALUES "
        f"({q(adj_no)},'RELOCATE',{q(relocation_trace[0])},(SELECT drug_id FROM drug_trace_inventory WHERE trace_code={q(relocation_trace[0])}),"
        f"(SELECT id FROM location_info WHERE location_code='OTC-A01-L1-P01'),(SELECT id FROM location_info WHERE location_code='OTC-A04-L1-P01'),"
        f"'IN_STOCK','IN_STOCK',{q('示例调拨：优化陈列货位')},(SELECT id FROM sys_user WHERE username='demo_warehouse'),"
        f"(SELECT id FROM inventory_task WHERE task_no={q(inv_task_no)}),'COMPLETED',{q(start + timedelta(hours=3))},{q(start + timedelta(hours=3))});"
    )
    sql.append(
        "INSERT INTO inventory_adjustment (adjust_no,adjust_type,trace_code,drug_id,before_status,after_status,reason,operator_id,related_task_id,status,created_at,updated_at) VALUES "
        f"({q('ADJ-' + DATE_STR + '-9002')},'LOSS',{q(lost_trace[0])},(SELECT drug_id FROM drug_trace_inventory WHERE trace_code={q(lost_trace[0])}),"
        f"'LOSS_CANDIDATE','LOST',{q('示例盘亏确认')},(SELECT id FROM sys_user WHERE username='demo_manager'),"
        f"(SELECT id FROM inventory_task WHERE task_no={q(inv_task_no)}),'COMPLETED',{q(start + timedelta(hours=3, minutes=10))},{q(start + timedelta(hours=3, minutes=10))});"
    )

    sql.extend(
        [
            "",
            "INSERT INTO report_export_task (task_id,report_type,export_format,query_params,status,file_id,message,requested_by,started_at,finished_at,created_at,updated_at) VALUES "
            f"({q('RPT-DEMO-' + DATE_STR + '-9001')},'SALES','xlsx','{{\"date\":\"2026-06-14\"}}','SUCCESS',NULL,{q('示例销售报表已生成')},"
            f"(SELECT id FROM sys_user WHERE username='demo_manager'),{q(ts(0, 18, 0))},{q(ts(0, 18, 1))},{q(ts(0, 18, 0))},{q(ts(0, 18, 1))});",
            "INSERT INTO audit_event (event_type,related_type,related_id,description,assigned_to,status,resolution,closed_at,severity,resolved_by,created_at,updated_at) VALUES "
            f"('LOW_STOCK','DRUG',{q('DEMO-' + DATE_STR + '-LOW-STOCK')},{q('示例低库存预警已处理')},(SELECT id FROM sys_user WHERE username='demo_manager'),1,"
            f"{q('已完成补货计划')},{q(ts(0, 17, 30))},'MEDIUM',(SELECT id FROM sys_user WHERE username='demo_manager'),{q(ts(0, 16, 0))},{q(ts(0, 17, 30))});",
            "INSERT INTO audit_event (event_type,related_type,related_id,description,assigned_to,status,resolution,closed_at,severity,ignored_by,ignored_at,created_at,updated_at) VALUES "
            f"('NEAR_EXPIRE','TRACE',{q('DEMO-' + DATE_STR + '-NEAR-EXPIRE')},{q('示例近效期预警已忽略')},(SELECT id FROM sys_user WHERE username='demo_manager'),2,"
            f"{q('演示数据无需处理')},{q(ts(0, 17, 40))},'LOW',(SELECT id FROM sys_user WHERE username='demo_manager'),{q(ts(0, 17, 40))},{q(ts(0, 16, 5))},{q(ts(0, 17, 40))});",
            "INSERT INTO notification (user_id,title,content,notification_type,business_type,business_id,read_at,created_at) VALUES "
            f"((SELECT id FROM sys_user WHERE username='demo_manager'),{q('示例报表任务完成')},{q('销售日报已生成，可在报表中心查看。')},'REPORT','REPORT_EXPORT',{q('DEMO-' + DATE_STR + '-REPORT')},{q(ts(0, 18, 5))},{q(ts(0, 18, 2))});",
            "INSERT INTO security_event (event_type,severity,user_id,username,ip,description,detail,handled,handled_by,handled_at,created_at) VALUES "
            f"('LOGIN_RISK','LOW',(SELECT id FROM sys_user WHERE username='demo_cashier'),'demo_cashier','127.0.0.1',{q('Codex demo seed: 示例安全事件已处理')},"
            f"'{{\"source\":\"demo-seed\"}}',TRUE,(SELECT id FROM sys_user WHERE username='demo_manager'),{q(ts(0, 18, 15))},{q(ts(0, 18, 10))});",
        ]
    )

    sql.extend(
        [
            "COMMIT;",
            "",
            "SELECT 'drug_info' AS table_name, COUNT(*) FROM drug_info WHERE remark LIKE '示例数据%'",
            "UNION ALL SELECT 'supplier', COUNT(*) FROM supplier WHERE remark='Codex demo seed supplier'",
            "UNION ALL SELECT 'location_info', COUNT(*) FROM location_info WHERE remark='Codex demo seed location'",
            "UNION ALL SELECT 'inbound_order', COUNT(*) FROM inbound_order WHERE order_no LIKE 'IN-20260614-9%'",
            "UNION ALL SELECT 'drug_trace_inventory', COUNT(*) FROM drug_trace_inventory WHERE trace_code LIKE '86000020260614%'",
            "UNION ALL SELECT 'sales_order', COUNT(*) FROM sales_order WHERE order_no LIKE 'SO-20260614-9%'",
            "UNION ALL SELECT 'audit_review', COUNT(*) FROM audit_review WHERE review_no LIKE 'REV-20260614-9%'",
            "UNION ALL SELECT 'trace_reservation', COUNT(*) FROM trace_reservation WHERE reservation_no LIKE 'RSV-20260614-9%'",
            "UNION ALL SELECT 'scan_task', COUNT(*) FROM scan_task WHERE task_no LIKE 'SCAN-20260614-9%'",
            "UNION ALL SELECT 'inventory_task', COUNT(*) FROM inventory_task WHERE task_no LIKE 'INV-20260614-9%'",
            "UNION ALL SELECT 'inventory_adjustment', COUNT(*) FROM inventory_adjustment WHERE adjust_no LIKE 'ADJ-20260614-9%';",
        ]
    )
    return "\n".join(sql)


def main():
    env = os.environ.copy()
    env["PGPASSWORD"] = "password_6KD3jP"
    proc = subprocess.run(
        [
            "psql",
            "-h",
            "127.0.0.1",
            "-U",
            "user_Q5e4RZ",
            "-d",
            "pharmacy_erp",
            "-v",
            "ON_ERROR_STOP=1",
        ],
        input=build_sql().encode("utf-8"),
        cwd=ROOT,
        env=env,
        capture_output=True,
    )
    print(proc.stdout.decode("utf-8", errors="replace"))
    if proc.returncode != 0:
        print(proc.stderr.decode("utf-8", errors="replace"))
        raise SystemExit(proc.returncode)


if __name__ == "__main__":
    main()
